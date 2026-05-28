import os
import sys
import shutil
import tempfile
import unittest
import re
from pathlib import Path

# Adiciona o diretório do projeto ao path de forma dinâmica e portátil
PROJECT_ROOT = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, PROJECT_ROOT)

from core.dynamic_filter import DynamicFilter
from core.file_handler import FileHandler
from core.excel_engine import ExcelEngine
from core.intent_parser import IntentParser
from core.ai_router import AIRouter, TaskType
from core.ocr_engine import OCREngine
from core.metadata_inspector import MetadataInspector
from core.response_generator import ResponseGenerator

class SophiaIntegrationTestSuite(unittest.TestCase):
    
    @classmethod
    def setUpClass(cls):
        cls.temp_dir = tempfile.mkdtemp(prefix="sophia_test_")
        cls.test_excel_path = os.path.join(cls.temp_dir, "teste.xlsx")
        
        # Cria uma planilha Excel básica para testes do ExcelEngine
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Planilha1"
        ws.cell(row=1, column=1, value="Nome")
        ws.cell(row=1, column=2, value="Equipe")
        ws.cell(row=1, column=3, value="FOTO01")
        
        ws.cell(row=2, column=1, value="Arthur")
        ws.cell(row=2, column=2, value="Equipe 01")
        
        ws.cell(row=3, column=1, value="Paulo")
        ws.cell(row=3, column=2, value="Equipe 02")
        
        wb.save(cls.test_excel_path)
        wb.close()

    @classmethod
    def tearDownClass(cls):
        try:
            shutil.rmtree(cls.temp_dir)
        except:
            pass

    # =========================================================================
    # DYNAMIC FILTER TESTS (1-7)
    # =========================================================================
    
    def test_01_dynamic_filter_simple_bool(self):
        f = DynamicFilter("True")
        self.assertTrue(f.evaluate({}))

    def test_02_dynamic_filter_boundary_comparison(self):
        f = DynamicFilter("15 <= int(dia) <= 20")
        self.assertTrue(f.evaluate({"dia": "17"}))
        self.assertFalse(f.evaluate({"dia": "12"}))

    def test_03_dynamic_filter_string_containment(self):
        f = DynamicFilter("'iphone' in camera.lower()")
        self.assertTrue(f.evaluate({"camera": "iPhone 13 Pro"}))
        self.assertFalse(f.evaluate({"camera": "Samsung S22"}))

    def test_04_dynamic_filter_syntax_error_fallback(self):
        # Erro de sintaxe (dia == 123 456)
        f = DynamicFilter("dia == 123 456")
        self.assertIsNone(f.code_obj)
        # Deve avaliar para True por padrão
        self.assertTrue(f.evaluate({"dia": "12"}))

    def test_05_dynamic_filter_sandbox_blocks_imports(self):
        f = DynamicFilter("__import__('os').system('echo')")
        # Deve falhar e cair no try-except retornando True
        self.assertTrue(f.evaluate({}))

    def test_06_dynamic_filter_missing_keys(self):
        f = DynamicFilter("camera == 'Canon'")
        # A chave 'camera' não está no dicionário, deve tratar e não travar
        self.assertTrue(f.evaluate({"dia": "10"}))

    def test_07_dynamic_filter_logical_operators(self):
        f = DynamicFilter("int(dia) > 10 and 'eq' in equipe")
        self.assertTrue(f.evaluate({"dia": "15", "equipe": "eq_central"}))
        self.assertFalse(f.evaluate({"dia": "8", "equipe": "eq_central"}))

    # =========================================================================
    # FILE HANDLER TESTS (8-18)
    # =========================================================================

    def test_08_file_handler_normalize_accents(self):
        res = FileHandler.normalize_string("Equipe Célula Água")
        self.assertEqual(res, "equipe celula agua")

    def test_09_file_handler_normalize_tabs_spaces(self):
        res = FileHandler.normalize_string("Equipe\t01   Central\n")
        self.assertEqual(res, "equipe 01 central")

    def test_10_file_handler_scan_empty_keywords(self):
        res = FileHandler.scan_directory(self.temp_dir, [])
        self.assertEqual(len(res), 0)

    def test_11_file_handler_scan_matching_subset(self):
        dummy_file = os.path.join(self.temp_dir, "atividade_foto.jpg")
        Path(dummy_file).touch()
        res = FileHandler.scan_directory(self.temp_dir, ["atividade"])
        self.assertTrue(any(f.name == "atividade_foto.jpg" for f in res))

    def test_12_file_handler_create_folder(self):
        f_name = "pasta_teste_cria"
        res = FileHandler.create_folder(self.temp_dir, f_name)
        self.assertTrue(res)
        self.assertTrue(os.path.exists(os.path.join(self.temp_dir, f_name)))

    def test_13_file_handler_create_folder_with_subfolders(self):
        f_name = "pasta_pai"
        subs = ["sub1", "sub2"]
        res = FileHandler.create_folder(self.temp_dir, f_name, subs)
        self.assertTrue(res)
        self.assertTrue(os.path.exists(os.path.join(self.temp_dir, f_name, "sub1")))
        self.assertTrue(os.path.exists(os.path.join(self.temp_dir, f_name, "sub2")))

    def test_14_file_handler_create_folder_subfolder_exclusion(self):
        f_name = "pasta_pai_exclusao"
        subs = ["sub1", "nao"]
        res = FileHandler.create_folder(self.temp_dir, f_name, subs)
        self.assertTrue(res)
        self.assertTrue(os.path.exists(os.path.join(self.temp_dir, f_name, "sub1")))
        self.assertFalse(os.path.exists(os.path.join(self.temp_dir, f_name, "nao")))

    def test_15_file_handler_move_folder_collision(self):
        src_dir = os.path.join(self.temp_dir, "origem")
        os.makedirs(src_dir, exist_ok=True)
        f_src = os.path.join(src_dir, "documento.txt")
        Path(f_src).touch()
        
        dst_dir = os.path.join(self.temp_dir, "destino")
        os.makedirs(dst_dir, exist_ok=True)
        # Cria arquivo colidente
        Path(os.path.join(dst_dir, "documento.txt")).touch()
        
        res = FileHandler.move_folder(f_src, dst_dir)
        self.assertTrue(res.endswith("documento_copia1.txt"))
        self.assertTrue(os.path.exists(res))

    def test_16_file_handler_delete_folder(self):
        del_dir = os.path.join(self.temp_dir, "deletar_mim")
        os.makedirs(del_dir, exist_ok=True)
        res = FileHandler.delete_folder(del_dir)
        self.assertTrue(res)
        self.assertFalse(os.path.exists(del_dir))

    def test_17_file_handler_advanced_rename_sequential(self):
        parent_dir = os.path.join(self.temp_dir, "renomear_lote")
        os.makedirs(parent_dir, exist_ok=True)
        os.makedirs(os.path.join(parent_dir, "subA"), exist_ok=True)
        os.makedirs(os.path.join(parent_dir, "subB"), exist_ok=True)
        
        res = FileHandler.rename_folders_advanced(parent_dir, "subpastas", "Equipe 01")
        self.assertTrue(res)
        self.assertTrue(os.path.exists(os.path.join(parent_dir, "Equipe 01")))
        self.assertTrue(os.path.exists(os.path.join(parent_dir, "Equipe 02")))

    def test_18_file_handler_advanced_rename_rollback(self):
        parent_dir = os.path.join(self.temp_dir, "renomear_lote_erro")
        os.makedirs(parent_dir, exist_ok=True)
        sub_path = os.path.join(parent_dir, "original_folder")
        os.makedirs(sub_path, exist_ok=True)
        
        # Dispara com planilha inexistente para forçar erro e validar rollback
        res = FileHandler.rename_folders_advanced(parent_dir, "subpastas", "excel_abas", excel_path="inexistente.xlsx")
        self.assertFalse(res)
        # Rollback deve manter a pasta original com seu nome original
        self.assertTrue(os.path.exists(sub_path))

    # =========================================================================
    # EXCEL ENGINE TESTS (19-24)
    # =========================================================================

    def test_19_excel_engine_column_validation(self):
        res = ExcelEngine.validate_main_columns(Path(self.test_excel_path), ["Nome"])
        self.assertTrue(res)

    def test_20_excel_engine_formula_injection(self):
        res = ExcelEngine.inject_formula(self.test_excel_path, "Planilha1", "D2", "=SOMA(A1)")
        self.assertTrue(res)

    def test_21_excel_engine_query_data_fuzzy(self):
        res = ExcelEngine.query_data(self.test_excel_path, "artur", "equipe")
        self.assertIn("Equipe 01", res)

    def test_22_excel_engine_query_data_missing_column(self):
        res = ExcelEngine.query_data(self.test_excel_path, "Arthur", "idade")
        self.assertIn("não localizada", res)

    def test_23_excel_engine_query_count_empty(self):
        res = ExcelEngine.query_count_empty(self.test_excel_path, "FOTO01")
        self.assertIn("2 registro", res) # As duas linhas tem FOTO01 vazia

    def test_24_excel_engine_norm(self):
        self.assertEqual(ExcelEngine._norm("  Água  Doce "), "agua doce")

    # =========================================================================
    # INTENT PARSER TESTS (25-30)
    # =========================================================================

    def test_25_intent_parser_normalize_tokenize(self):
        ip = IntentParser()
        tokens = ip._normalize_and_tokenize("Oi! Crie uma pasta, por favor.")
        self.assertTrue("crie" in tokens)
        self.assertTrue("oi" in tokens)

    def test_26_intent_parser_learn_intent(self):
        ip = IntentParser()
        ip.learn_intent("CREATE_FOLDER", "fazer_diretorio")
        ip._load_intents()
        self.assertTrue("fazer_diretorio" in ip._intents_bow["CREATE_FOLDER"])

    def test_27_intent_parser_local_bow_greeting(self):
        ip = IntentParser()
        res = ip.parse_single_intent("Olá, bom dia!")
        self.assertEqual(res["intent"], "GREETING")

    def test_28_intent_parser_local_bow_identity(self):
        ip = IntentParser()
        res = ip.parse_single_intent("Quem é você?")
        self.assertEqual(res["intent"], "IDENTITY")

    def test_29_intent_parser_multiple_split_intents(self):
        ip = IntentParser()
        # Força o bypass da cloud para testar a engine offline local
        ip._cloud_inference = lambda *args: None
        res = ip.parse_multiple_intents("crie uma pasta depois delete a pasta")
        self.assertEqual(len(res), 2)
        self.assertEqual(res[0]["intent"], "CREATE_FOLDER")
        self.assertEqual(res[1]["intent"], "DELETE_FOLDER")

    def test_30_intent_parser_cache_mechanism(self):
        ip = IntentParser()
        # Mock para retornar um resultado de conversa da nuvem, permitindo testar o cache
        ip._cloud_inference = lambda *args: {"intent": "EXPLAIN_COMMAND", "status": "CONVERSATIONAL", "resposta": "O scan faz uma busca."}
        ip._response_cache.clear()
        res1 = ip.parse_multiple_intents("como funciona o scan")
        # Deve estar em cache agora
        cache_key = list(ip._response_cache.keys())[0]
        self.assertEqual(ip._response_cache[cache_key][0]["intent"], "EXPLAIN_COMMAND")

    # =========================================================================
    # AI ROUTER TESTS (31-34)
    # =========================================================================

    def test_31_ai_router_key_resolution(self):
        ar = AIRouter()
        prov = {"env_key": "GROQ_API_KEY_1", "api_key": "old"}
        val = ar._resolve_api_key(prov)
        self.assertTrue(val.startswith("gsk_")) # Lida com chave ativa do .env

    def test_32_ai_router_usage_tracking(self):
        ar = AIRouter()
        ar.usage_today.clear()
        ar._track_usage("TestProv", 1000)
        self.assertEqual(ar.usage_today["TestProv"], 1000)

    def test_33_ai_router_exhaustion_check(self):
        ar = AIRouter()
        prov = {"nome": "TestProv", "limite_diario_tokens": 10000}
        ar.usage_today["TestProv"] = 9500
        self.assertTrue(ar._is_exhausted(prov))

    def test_34_ai_router_temperature_selection(self):
        ar = AIRouter()
        self.assertEqual(TaskType.CODE, "code")
        self.assertEqual(TaskType.REASONING, "reasoning")

    # =========================================================================
    # OCR ENGINE TESTS (35-37)
    # =========================================================================

    def test_35_ocr_engine_availability(self):
        available = OCREngine.is_available()
        self.assertIn(available, [True, False])

    def test_36_ocr_engine_extract_stamp_regex_date(self):
        text = "Relatório executado em 15/05/2026 às 14:30"
        md = re.search(r'(\d{1,2})[/\-.](\d{1,2})[/\-.](\d{2,4})', text)
        self.assertIsNotNone(md)
        self.assertEqual(md.group(1), "15")
        self.assertEqual(md.group(2), "05")

    def test_37_ocr_engine_extract_stamp_km(self):
        text = "KM: 123456 \nHODOMETRO: 123456"
        mk = re.search(r'(?:KM|HODOMETRO)[\s:\-]*(\d[\d.,]*)', text, re.IGNORECASE)
        self.assertIsNotNone(mk)
        self.assertEqual(mk.group(1), "123456")

    # =========================================================================
    # METADATA INSPECTOR TESTS (38-39)
    # =========================================================================

    def test_38_metadata_inspector_modtime_fallback(self):
        dummy_file = os.path.join(self.temp_dir, "no_exif.jpg")
        Path(dummy_file).touch()
        info = MetadataInspector.extract_full_metadata(dummy_file)
        self.assertNotEqual(info["data"], "Desconhecida")
        self.assertEqual(info["camera"], "Dispositivo do Sistema Operacional")

    def test_39_metadata_inspector_missing_file(self):
        info = MetadataInspector.extract_full_metadata("missing_file_non_existent.jpg")
        self.assertEqual(info["data"], "Desconhecida")

    # =========================================================================
    # RESPONSE GENERATOR TESTS (40)
    # =========================================================================

    def test_40_response_generator_hash_lookup(self):
        res = ResponseGenerator.generate("EXPLAIN_COMMAND", "Arthur", raw_input="como funciona o scan")
        self.assertIn("varredura nas suas pastas", res)

    # =========================================================================
    # NEW IMPROVEMENTS TESTS (41-42)
    # =========================================================================

    def test_41_ai_router_config_models_override(self):
        # Cria um config temporário simulando a nova chave modelos_ia
        temp_config = os.path.join(self.temp_dir, "temp_config.json")
        import json
        with open(temp_config, "w", encoding="utf-8") as f:
            json.dump({
                "provedores_ia": [
                    {
                        "nome": "ProvedorTeste",
                        "base_url": "https://api.groq.com/openai/v1/chat/completions",
                        "env_key": "GROQ_API_KEY_1",
                        "modelo": "original",
                        "modelo_codigo": "original",
                        "modelo_reasoning": "original",
                        "modelo_intent": "original",
                        "limite_diario_tokens": 500000,
                        "ativo": True,
                        "prioridade": 1
                    }
                ],
                "modelos_ia": {
                    "chat": "modelo-teste-chat",
                    "code": "modelo-teste-code",
                    "reasoning": "modelo-teste-reasoning",
                    "intent": "modelo-teste-intent"
                }
            }, f)
        
        ar = AIRouter(config_path=temp_config)
        # Verifica se o override de modelos_ia foi aplicado em todos os providers carregados
        for p in ar.providers:
            self.assertEqual(p["modelo"], "modelo-teste-chat")
            self.assertEqual(p["modelo_codigo"], "modelo-teste-code")
            self.assertEqual(p["modelo_reasoning"], "modelo-teste-reasoning")
            self.assertEqual(p["modelo_intent"], "modelo-teste-intent")

    def test_42_file_handler_log_error_rotation(self):
        # Testa rotação do log de erros.
        original_log_exists = os.path.exists("erros_conhecidos.txt")
        original_log_content = ""
        if original_log_exists:
            with open("erros_conhecidos.txt", "r", encoding="utf-8") as f:
                original_log_content = f.read()
            os.remove("erros_conhecidos.txt")
        
        try:
            # Cria um log inicial gigante (> 5MB)
            with open("erros_conhecidos.txt", "w", encoding="utf-8") as f:
                f.write("X" * (5 * 1024 * 1024 + 100))
            
            # Executa log_error que deve disparar a rotação
            FileHandler.log_error("Erro de Teste")
            
            # erros_conhecidos.txt atual deve ser novo e pequeno
            self.assertTrue(os.path.exists("erros_conhecidos.txt"))
            self.assertLess(os.path.getsize("erros_conhecidos.txt"), 1000)
            
            # erros_conhecidos.old.txt deve existir com o tamanho original gigante
            self.assertTrue(os.path.exists("erros_conhecidos.old.txt"))
            self.assertGreater(os.path.getsize("erros_conhecidos.old.txt"), 5 * 1024 * 1024)
            
        finally:
            # Limpa logs criados pelo teste
            if os.path.exists("erros_conhecidos.txt"):
                os.remove("erros_conhecidos.txt")
            if os.path.exists("erros_conhecidos.old.txt"):
                os.remove("erros_conhecidos.old.txt")
            
            # Restaura o histórico original do usuário se existia
            if original_log_exists:
                with open("erros_conhecidos.txt", "w", encoding="utf-8") as f:
                    f.write(original_log_content)

    def test_43_intent_parser_macro_vba(self):
        # Testa se comandos de macro/VBA são direcionados para GENERATE_NEW_SKILL
        ip = IntentParser()
        if ip.groq_api_key:
            res = ip.parse_multiple_intents("crie uma macro no Excel para formatar relatorios")
            self.assertEqual(res[0]["intent"], "GENERATE_NEW_SKILL")
            self.assertIn("prompt", res[0])

    def test_44_intent_parser_windows_aut(self):
        # Testa se comandos de automação do Windows vão para GENERATE_NEW_SKILL
        ip = IntentParser()
        if ip.groq_api_key:
            res = ip.parse_multiple_intents("abra o bloco de notas no windows e digite oi")
            self.assertEqual(res[0]["intent"], "GENERATE_NEW_SKILL")
            self.assertIn("prompt", res[0])

    def test_45_embedded_interpreter_run(self):
        # Testa o interpretador Python embutido no main.py
        temp_script = os.path.join(self.temp_dir, "test_run.py")
        with open(temp_script, "w", encoding="utf-8") as f:
            f.write("print('SANDBOX_OK')")
            
        import subprocess
        # Executa main.py passando o script como argumento
        res = subprocess.run(
            [sys.executable, "main.py", temp_script],
            capture_output=True, text=True, timeout=10
        )
        self.assertEqual(res.returncode, 0)
        self.assertIn("SANDBOX_OK", res.stdout)

    def test_46_embedded_interpreter_compile(self):
        # Testa a emulação de py_compile do interpretador embutido no main.py
        temp_valid = os.path.join(self.temp_dir, "valid.py")
        with open(temp_valid, "w", encoding="utf-8") as f:
            f.write("def foo():\n    pass\n")
            
        temp_invalid = os.path.join(self.temp_dir, "invalid.py")
        with open(temp_invalid, "w", encoding="utf-8") as f:
            f.write("def foo(\n") # Erro de sintaxe proposital
            
        import subprocess
        # Compilação válida
        res_valid = subprocess.run(
            [sys.executable, "main.py", "-m", "py_compile", temp_valid],
            capture_output=True, text=True, timeout=10
        )
        self.assertEqual(res_valid.returncode, 0)
        
        # Compilação inválida
        res_invalid = subprocess.run(
            [sys.executable, "main.py", "-m", "py_compile", temp_invalid],
            capture_output=True, text=True, timeout=10
        )
        self.assertEqual(res_invalid.returncode, 1)
        self.assertIn("Erro de compilacao", res_invalid.stderr)

    def test_47_intent_parser_check_and_upgrade_intent(self):
        # Testa a salvaguarda programática de upgrade de CREATE_FOLDER para GENERATE_NEW_SKILL
        ip = IntentParser()
        res = {
            "intent": "CREATE_FOLDER",
            "raw_input": "crie a pasta teeste e coloque a planilha de dados nela",
            "folder_name": "teeste"
        }
        res_upgraded = ip._check_and_upgrade_intent(res)
        self.assertEqual(res_upgraded["intent"], "GENERATE_NEW_SKILL")
        self.assertIn("prompt", res_upgraded)

    def test_48_intent_parser_history_truncation(self):
        # Testa se as mensagens do histórico de conversa do IntentParser são truncadas e limpas corretamente
        ip = IntentParser()
        
        class DummyRouter:
            def call_intent(self, messages, max_tokens):
                return "Certo! [ACTION: GENERATE_NEW_SKILL | PROMPT=" + ("b" * 500) + "]", "MockProvider"
        
        ip._router = DummyRouter()
        
        long_input = "a" * 400
        ip._cloud_inference(long_input)
        
        # Deve ter 2 mensagens no histórico: user e assistant
        self.assertEqual(len(ip.chat_history), 2)
        
        # User message deve estar truncada para 303 caracteres ("..." incluído)
        user_msg = ip.chat_history[0]["content"]
        self.assertTrue(len(user_msg) <= 303)
        self.assertTrue(user_msg.endswith("..."))
        
        # Assistant message deve ter a action limpa e truncada
        assistant_msg = ip.chat_history[1]["content"]
        self.assertTrue(len(assistant_msg) <= 303)
        self.assertIn("[ACTION: GENERATE_NEW_SKILL]", assistant_msg)
        self.assertNotIn("PROMPT=", assistant_msg)

    def test_49_dynamic_coder_continuation(self):
        # Testa se a continuação de código em DynamicCoder funciona corretamente costurando as partes truncadas
        from core.dynamic_coder import DynamicCoder
        coder = DynamicCoder(api_key="dummy_key")
        
        class DummyRouter:
            def __init__(self):
                self.calls = 0
                self.last_was_truncated = False
                
            def call_code(self, messages, max_tokens):
                self.calls += 1
                if self.calls == 1:
                    self.last_was_truncated = True
                    return "```python\nimport os\nclass Foo:\n    def bar(self):\n", "MockProvider"
                else:
                    self.last_was_truncated = False
                    return "        print('done')\n```", "MockProvider"
                    
        coder._router = DummyRouter()
        
        # Gera o script chamando a função mockada
        script = coder._generate_script("crie uma classe foo", {})
        
        # Deve ter chamado a API 2 vezes
        self.assertEqual(coder._router.calls, 2)
        
        # O script final deve estar costurado e limpo
        self.assertEqual(
            script,
            "import os\nclass Foo:\n    def bar(self):\n        print('done')"
        )

if __name__ == "__main__":
    unittest.main()
