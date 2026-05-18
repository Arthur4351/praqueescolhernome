import os
import glob
from pathlib import Path
import openpyxl
from PIL import Image, ImageDraw, ImageFont

def criar_imagem(texto, caminho_arquivo, cor_fundo):
    img = Image.new('RGB', (800, 600), color=cor_fundo)
    d = ImageDraw.Draw(img)
    try:
        font = ImageFont.truetype("arial.ttf", 40)
    except:
        font = ImageFont.load_default()
    
    try:
        bbox = d.textbbox((0, 0), texto, font=font)
        text_width = bbox[2] - bbox[0]
        text_height = bbox[3] - bbox[1]
    except AttributeError:
        text_width, text_height = d.textsize(texto, font=font)
        
    x = (800 - text_width) / 2
    y = (600 - text_height) / 2
    d.text((x, y), texto, fill=(255, 255, 255), font=font)
    img.save(caminho_arquivo)

def encontrar_excel():
    base_dir = os.path.dirname(__file__)
    alvo = "03 - Relatório Fotográfico Conserva PR Vias (1)"
    for root, dirs, files in os.walk(base_dir):
        for file in files:
            if alvo in file and not file.startswith('~'):
                return os.path.join(root, file)
                
    for root, dirs, files in os.walk(base_dir):
        for file in files:
            if file.endswith(('.xlsm', '.xlsx')) and not file.startswith('~'):
                if 'Relatório Fotográfico' in file:
                    return os.path.join(root, file)
    return None

def gerar_teste():
    excel_path = encontrar_excel()
    if not excel_path:
        print("ERRO: Nenhuma planilha de Relatório Fotográfico encontrada no projeto.")
        return
        
    print(f"Usando planilha: {os.path.basename(excel_path)}")
    
    try:
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True, keep_vba=True)
    except:
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
        
    abas = wb.sheetnames
    wb.close()
    
    equipes = [aba.strip() for aba in abas if aba.lower().strip() not in ('resumo', 'dashboard', 'base', 'config', 'plan1', 'parametros gerais')]
    print(f"Abas encontradas (Equipes): {equipes}")
    
    pasta_teste = os.path.join(os.path.dirname(__file__), "fotos_teste_por_aba")
    if not os.path.exists(pasta_teste):
        os.makedirs(pasta_teste)
        
    tipos = [
        ("E", "Entrada", (50, 150, 50)),
        ("A", "Antes", (150, 50, 50)),
        ("D", "Depois", (50, 50, 150)),
        ("S", "Saida", (100, 100, 100))
    ]
    
    total = 0
    
    for equipe in equipes:
        pasta_equipe = os.path.join(pasta_teste, equipe)
        if not os.path.exists(pasta_equipe):
            os.makedirs(pasta_equipe)
            
        for dia in range(1, 11): # Dia 1 ao 10
            dia_str = f"{dia:02d}" # "01", "02", etc.
            
            for sufixo, fase_nome, cor in tipos:
                # Exemplo: "01 1-E.jpg"
                nome_arquivo = f"{dia_str} {sufixo}.jpg"
                caminho = os.path.join(pasta_equipe, nome_arquivo)
                
                texto = f"Equipe: {equipe}\nDia: {dia_str}\nFase: {fase_nome}"
                criar_imagem(texto, caminho, cor)
                total += 1
                
    print(f"SUCESSO: Geradas {total} fotos distribuidas em {len(equipes)} pastas no diretorio: fotos_teste_por_aba")

if __name__ == "__main__":
    gerar_teste()
