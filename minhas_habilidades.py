import os, re, gc, datetime, uuid, io, concurrent.futures, difflib, unicodedata
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from PIL import Image as PILImage, ImageOps
from core.metadata_inspector import MetadataInspector
from core.ocr_engine import OCREngine
import pandas as pd
from core.file_handler import FileHandler

def remover_acentos(texto):
    if not texto: return ""
    return "".join(c for c in unicodedata.normalize('NFD', str(texto)) if unicodedata.category(c) != 'Mn').upper()

class UnclosableBytesIO(io.BytesIO):
    def close(self): pass

class SophiaExecutor:
    def __init__(self, autonomous_mode=False):
        self._ocr_disponivel = OCREngine.is_available()
        self.autonomous_mode = autonomous_mode

    def extrair_dia(self, valor):
        if isinstance(valor, datetime.datetime): return str(valor.day)
        m = re.search(r'(\d{1,2})', str(valor))
        return str(int(m.group(1))) if m else None

    def preparar_foto(self, caminho_foto, extrair_ocr=False):
        try:
            selo = None
            if extrair_ocr and self._ocr_disponivel:
                with open(caminho_foto, "rb") as f:
                    selo = OCREngine.extract_stamp_from_bytes(f.read())
            
            img_io = UnclosableBytesIO()
            with PILImage.open(caminho_foto) as img:
                img = ImageOps.exif_transpose(img)
                # Reduz a imagem para o Excel e salva como JPEG comprimido para não explodir os 4GB de RAM
                img.thumbnail((600, 600))
                if img.mode != 'RGB':
                    img = img.convert('RGB')
                img.save(img_io, format="JPEG", quality=75)
            
            img_io.seek(0)
            gc.collect()
            return img_io, selo
        except Exception as e:
            with open("erros_conhecidos.txt", "a", encoding="utf-8") as err_log:
                err_log.write(f"Erro em preparar_foto ({caminho_foto}): {e}\n")
            return None, None

    def atualizar_datas_planilha(self, excel_path, pasta_fotos, log_func):
        """Preenche sequencialmente de 1 a 30 nas colunas de data/dia, sem ler imagens."""
        try:
            from openpyxl import load_workbook
            wb = load_workbook(excel_path, keep_vba=True)
            total_corrigido = 0
            
            for nome_aba in wb.sheetnames:
                ws = wb[nome_aba]
                coluna_data = None
                row_cabecalho = 1
                
                # Busca nas primeiras 5 linhas por um cabeçalho que contenha DATA ou DIA
                for r in range(1, 6):
                    for c in range(1, ws.max_column + 1):
                        v = ws.cell(row=r, column=c).value
                        if v and ('DATA' in str(v).upper() or 'DIA' in str(v).upper()):
                            coluna_data = c
                            row_cabecalho = r
                            break
                    if coluna_data: break
                
                if coluna_data:
                    dia = 1
                    for r in range(row_cabecalho + 1, ws.max_row + 1):
                        # Só preenche se a linha tiver algum conteúdo (para não preencher o infinito)
                        linha_vazia = True
                        for check_c in range(1, ws.max_column + 1):
                            if ws.cell(row=r, column=check_c).value is not None:
                                linha_vazia = False
                                break
                        
                        if not linha_vazia:
                            ws.cell(row=r, column=coluna_data).value = dia
                            dia += 1
                            if dia > 30:
                                dia = 1 # Reinicia a contagem
                            total_corrigido += 1
            
            wb.save(excel_path)
            log_func(f"✅ Datas preenchidas sequencialmente (1-30): {total_corrigido} célula(s).")
            import gc
            gc.collect()
        except Exception as e:
            log_func(f"❌ Erro ao atualizar datas: {e}")

    def _injetar_foto_segura(self, ws, img_io, r, c):
        if not c: return 0
        
        try:
            from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor, AnchorMarker
            img = ExcelImage(img_io)
            
            # Detecta se a célula alvo faz parte de uma mesclagem (Merged Cells)
            alvo_coord = ws.cell(row=r, column=c).coordinate
            range_mesclado = None
            for merged_range in ws.merged_cells.ranges:
                if alvo_coord in merged_range:
                    range_mesclado = merged_range
                    break
            
            # Define os limites da âncora (Canto Superior Esquerdo e Inferior Direito)
            if range_mesclado:
                min_col, min_row, max_col, max_row = range_mesclado.bounds
                # openpyxl bounds são 1-based. AnchorMarker usa 0-based.
                m1 = AnchorMarker(col=min_col-1, colOff=0, row=min_row-1, rowOff=0)
                m2 = AnchorMarker(col=max_col, colOff=0, row=max_row, rowOff=0)
            else:
                m1 = AnchorMarker(col=c-1, colOff=0, row=r-1, rowOff=0)
                m2 = AnchorMarker(col=c, colOff=0, row=r, rowOff=0)
            
            img.anchor = TwoCellAnchor(_from=m1, to=m2)
            ws.add_image(img)
            
            gc.collect()
            return 1
        except Exception as e:
            # Fallback seguro caso a ancoragem complexa falhe
            try:
                ws.add_image(ExcelImage(img_io), ws.cell(row=r, column=c).coordinate)
                return 1
            except:
                return 0

    def _injetar_km_dinamico(self, ws, r_base, selo, tipo):
        if not selo or not selo.get('km'): return
        km_cols = []
        for r_atual in range(max(1, r_base - 2), r_base + 6):
            for c in range(1, ws.max_column + 1):
                v = ws.cell(row=r_atual, column=c).value
                if not v: continue
                vn = str(v).upper().replace(" ", "")
                if 'KM' in vn or 'HODOMETRO' in vn:
                    km_cols.append((r_atual, c))
        if not km_cols: return
        km_cols.sort(key=lambda x: (x[0], x[1]))
        t_cell = km_cols[0] if tipo == 'INICIAL' or len(km_cols) == 1 else km_cols[1]
        try:
            ws.cell(row=t_cell[0], column=t_cell[1] + 1).value = int(selo['km'])
        except:
            ws.cell(row=t_cell[0], column=t_cell[1] + 1).value = selo['km']

    def _coletar_fotos_pasta(self, pasta):
        """Coleta metadados das fotos de uma pasta sem carregar os pixels ainda."""
        resultado = {}
        eq = pasta.name.strip().upper()
        
        # Carrega padrões ensinados dinamicamente (NLP)
        config_path = Path(__file__).parent / "config.json"
        padroes = {}
        if config_path.exists():
            try:
                import json
                with open(config_path, "r", encoding="utf-8") as f:
                    cfg = json.load(f)
                    padroes = cfg.get("padroes_fotos", {})
            except: pass

        for f in list(pasta.rglob("*.[jJ][pP]*[gG]")) + list(pasta.rglob("*.[pP][nN][gG]")):
            nm = f.stem.upper()
            m = re.search(r'(\d+)', nm)
            if not m: continue
            dia = str(int(m.group(1)))
            if dia not in resultado:
                resultado[dia] = {}
                
            t = None
            if padroes:
                sufixo_foto = nm
                if nm.startswith(dia):
                    sufixo_foto = nm[len(dia):].lstrip(" -_")
                elif nm.endswith(dia):
                    sufixo_foto = nm[:-len(dia)].rstrip(" -_")
                
                for chave, destino in padroes.items():
                    chave_upper = chave.upper()
                    m_ch = re.search(r'(\d+)', chave_upper)
                    if m_ch:
                        cd = m_ch.group(1)
                        sufixo_chave = chave_upper
                        if chave_upper.startswith(cd):
                            sufixo_chave = chave_upper[len(cd):].lstrip(" -_")
                        elif chave_upper.endswith(cd):
                            sufixo_chave = chave_upper[:-len(cd)].rstrip(" -_")
                    else:
                        sufixo_chave = chave_upper.lstrip(" -_")
                    
                    if sufixo_foto == sufixo_chave:
                        t = str(destino).upper().strip()
                        break
                        
            if not t:
                t = (
                    'ENTRADA' if re.search(r'\bE\b|\bENTRADA\b', nm) else
                    'SAIDA'   if re.search(r'\bS\b|\bSAIDA\b', nm) else
                    'ANTES'   if re.search(r'\bA\b|\(A\)|\bANTES\b', nm) else
                    'DEPOIS'  if re.search(r'\bD\b|\(D\)|\bDEPOIS\b|\(2\)|\+', nm) else
                    None
                )
            if not t:
                if 'ANTES' not in resultado[dia]:
                    t = 'ANTES'
                elif 'ENTRADA' not in resultado[dia]:
                    t = 'ENTRADA'
                elif 'DEPOIS' not in resultado[dia]:
                    t = 'DEPOIS'
                else:
                    t = 'SAIDA'
            t = remover_acentos(t)
            resultado[dia][t] = {'path': f, 'ocr': t in ['ENTRADA', 'SAIDA']}
        return eq, resultado

    def _processar_fotos_lote(self, dias_dict, max_workers):
        """Carrega e redimensiona apenas as fotos de um lote (uma equipe/aba)."""
        tasks = []
        for dia, tipos in dias_dict.items():
            for tipo, meta in tipos.items():
                if 'path' in meta:
                    tasks.append((dia, tipo, meta['path'], meta.get('ocr', False)))

        with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as pool:
            future_map = {
                pool.submit(self.preparar_foto, tk[2], tk[3]): tk
                for tk in tasks
            }
            for future in concurrent.futures.as_completed(future_map):
                dia, tipo, _, _ = future_map[future]
                try:
                    img_io, selo = future.result()
                    if img_io:
                        dias_dict[dia][tipo]['p'] = img_io
                        dias_dict[dia][tipo]['s'] = selo
                except:
                    pass

        return dias_dict

    def _injetar_aba(self, ws, aba_norm, dias_cache, dados_relatorio):
        """Injeta fotos em uma única aba do Excel e retorna total injetado."""
        fotos_total = 0
        dia_atual = None
        max_r = ws.max_row + 1
        max_c = min(ws.max_column + 1, 200)
        linhas_em_branco = 0
        dias_na_planilha = set()

        for r in range(1, max_r):
            linha_tem_dados = False
            for c in range(1, max_c):
                v = ws.cell(row=r, column=c).value
                if not v: continue
                linha_tem_dados = True
                v_norm = remover_acentos(v).replace(" ", "").replace(".", "")

                if 'DATA' in v_norm or 'DIA' in v_norm:
                    for off in range(1, 6):
                        if c + off >= max_c: break
                        v_off = ws.cell(row=r, column=c + off).value
                        if v_off:
                            d = self.extrair_dia(v_off)
                            if d:
                                dia_atual = d
                                dias_na_planilha.add(dia_atual)
                                break

                if dia_atual and dia_atual in dias_cache:
                    fots = dias_cache[dia_atual]
                    if not fots.get('_verificado'):
                        fots['_verificado'] = True

                    # Mapeamento dinâmico: cruza os destinos mapeados com o título da coluna
                    for destino, meta_foto in list(fots.items()):
                        if destino.startswith('_'): continue
                        
                        destino_norm = remover_acentos(destino)
                        match_destino = False
                        if destino_norm == 'ENTRADA' and any(m in v_norm for m in ['FOTO01', 'ENTRADA', 'FOTODEENTRADA', '-E']): match_destino = True
                        elif destino_norm == 'SAIDA' and any(m in v_norm for m in ['FOTO02', 'SAIDA', 'FOTODESAIDA', '-S']): match_destino = True
                        elif destino_norm == 'ANTES' and any(m in v_norm for m in ['ANTES', 'SERVICO01', 'FOTODEANTES', 'ATIVIDADEREALISADA', 'ATIVIDADEREALIZADA', '-A']): match_destino = True
                        elif destino_norm == 'DEPOIS' and any(m in v_norm for m in ['DEPOIS', 'SERVICO02', 'FOTODEDEPOIS', '-D']): match_destino = True
                        elif destino_norm in v_norm: match_destino = True
                        elif difflib.SequenceMatcher(None, destino_norm, v_norm).ratio() > 0.8: match_destino = True
                            
                        if match_destino and 'p' in meta_foto and not meta_foto.get('_injetada'):
                            meta_foto['p'].seek(0)
                            fotos_total += self._injetar_foto_segura(ws, meta_foto['p'], r + 1, c)
                            meta_foto['_injetada'] = True
                            
                            tipo_km = 'INICIAL' if destino in ['ANTES', 'ENTRADA'] else 'FINAL'
                            self._injetar_km_dinamico(ws, r, meta_foto.get('s'), tipo_km)

            if not linha_tem_dados:
                linhas_em_branco += 1
                if linhas_em_branco >= 50:
                    break
            else:
                linhas_em_branco = 0

        if aba_norm not in dados_relatorio:
            dados_relatorio[aba_norm] = {}
            
        for dia in sorted(list(dias_na_planilha), key=lambda x: int(x)):
            fots = dias_cache.get(dia, {})
            faltas = [tp for tp in ['ENTRADA', 'SAIDA', 'ANTES', 'DEPOIS'] if tp not in fots or 'p' not in fots[tp]]
            if faltas:
                dados_relatorio[aba_norm][dia] = faltas

        return fotos_total

    def processar_comando(self, pasta_raiz_fotos, excel_path, destino_path, nome_usuario, log_func):
        """
        Motor principal de injeção de fotos.
        Arquitetura em lotes por aba — processa e libera memória aba a aba
        para suportar volumes de 2000+ fotos sem travamento.
        """
        try:
            destino_path = Path(destino_path)
            if not destino_path.exists():
                destino_path.mkdir(parents=True)

            max_workers = min(4, os.cpu_count() or 1)

            # Monta índice leve: {NOME_EQUIPE: Path(pasta)} sem carregar pixels
            indice_pastas = {}
            for pasta in [d for d in Path(pasta_raiz_fotos).iterdir() if d.is_dir()]:
                eq_norm = pasta.name.strip().upper()
                if eq_norm in indice_pastas:
                    log_func(f"⚠️ Atenção: duas pastas normalizam para o mesmo nome '<b>{eq_norm}</b>'. A segunda será ignorada.")
                else:
                    indice_pastas[eq_norm] = pasta

            wb = None
            for attempt in range(3):
                try:
                    wb = load_workbook(excel_path, keep_vba=True)
                    break
                except PermissionError:
                    if attempt < 2:
                        log_func(f"⚠️ O Excel original ('{Path(excel_path).name}') está aberto! Salve e feche-o. Vou tentar ler de novo em 10 segundos... (Tentativa {attempt+1}/3)")
                        import time
                        time.sleep(10)
                    else:
                        log_func("❌ ERRO CRÍTICO: Não consegui abrir o Excel porque ele continua bloqueado. Abortando.")
                        return "Abortado: Excel bloqueado."
            
            fotos_total_geral = 0
            dados_relatorio = {}
            streams_manter_vivos = []

            for nm_aba in wb.sheetnames:
                aba_norm = nm_aba.strip().upper()

                # Fuzzy match: encontra a pasta que corresponde a esta aba
                pasta_alvo = indice_pastas.get(aba_norm)
                if pasta_alvo is None:
                    for eq_key, eq_pasta in indice_pastas.items():
                        ratio = difflib.SequenceMatcher(None, aba_norm, eq_key).ratio()
                        if ratio > 0.75 or aba_norm in eq_key or eq_key in aba_norm:
                            pasta_alvo = eq_pasta
                            break

                if pasta_alvo is None:
                    log_func(f"⚠️ Aba '{nm_aba}': nenhuma pasta correspondente encontrada. Pulando.")
                    continue

                log_func(f"⏳ Processando aba <b>{nm_aba}</b>...")

                # Coleta metadados (sem carregar pixels)
                _, dias_meta = self._coletar_fotos_pasta(pasta_alvo)

                if not dias_meta:
                    log_func(f"ℹ️ Aba '{nm_aba}': pasta sem fotos reconhecíveis.")
                    continue

                # Carrega pixels apenas desta aba em paralelo
                dias_carregado = self._processar_fotos_lote(dias_meta, max_workers)

                # Injeta na aba
                ws = wb[nm_aba]
                fotos_aba = self._injetar_aba(ws, aba_norm, dias_carregado, dados_relatorio)
                fotos_total_geral += fotos_aba

                log_func(f"✅ Aba <b>{nm_aba}</b>: {fotos_aba} foto(s) injetada(s).")

                # Guarda referências dos streams para o openpyxl conseguir salvar no final
                for dia_data in dias_carregado.values():
                    for tipo_data in dia_data.values():
                        if isinstance(tipo_data, dict) and 'p' in tipo_data:
                            streams_manter_vivos.append(tipo_data['p'])

                del dias_carregado, dias_meta
                gc.collect()

            # (Log de divergências removido daqui para focar no relatório final formatado)

            if fotos_total_geral > 0:
                try:
                    nome_saida = destino_path / f"Relatorio_{uuid.uuid4().hex[:4]}.xlsm"
                    wb.save(nome_saida)
                    
                    # Agora podemos liberar os streams
                    for s in streams_manter_vivos:
                        try: s.close()
                        except: pass
                    del streams_manter_vivos
                    
                    log_func(f"🎉 <b>CONCLUÍDO!</b> {fotos_total_geral} fotos injetadas → <b>{nome_saida.name}</b>")
                    
                    # Ordem do Usuário: Gerar relatório automático ao fim
                    try:
                        from core.relatorio_engine import gerar_relatorio_faltas
                        import json
                        
                        log_func("📊 <b>SOPHIA:</b> Gerando relatório de faltas automático...")
                        resumo, dados = gerar_relatorio_faltas(dados_relatorio)
                        log_func(resumo)
                        
                        # Salvar JSON para o n8n
                        rel_dir = Path("relatorios_n8n")
                        rel_dir.mkdir(exist_ok=True)
                        json_path = rel_dir / f"relatorio_faltas_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
                        with open(json_path, 'w', encoding='utf-8') as f:
                            json.dump(dados, f, indent=4, ensure_ascii=False)
                            
                        log_func(f"💾 JSON salvo em: {json_path.absolute()}")
                        
                        # Ordem do Usuário: Enviar para WhatsApp via CallMeBot
                        try:
                            import requests
                            from urllib.parse import quote
                            
                            log_func("🌐 <b>SOPHIA:</b> Enviando relatório direto para o seu WhatsApp via CallMeBot...")
                            
                            telefone = "557399310836"
                            apikey = "4667834"
                            config_path = Path(__file__).parent / "config.json"
                            if config_path.exists():
                                try:
                                    import json
                                    with open(config_path, "r", encoding="utf-8") as fc:
                                        cfg = json.load(fc)
                                        telefone = cfg.get("whatsapp_telefone", telefone)
                                        apikey = cfg.get("whatsapp_apikey", apikey)
                                except Exception: pass
                                
                            texto_codificado = quote(resumo)
                            
                            url_callmebot = f"https://api.callmebot.com/whatsapp.php?phone={telefone}&text={texto_codificado}&apikey={apikey}"
                            
                            resp = requests.get(url_callmebot, timeout=10)
                            if resp.status_code in [200, 201]:
                                log_func("✅ Relatório enviado com sucesso para o seu celular!")
                            else:
                                log_func(f"⚠️ Falha ao enviar WhatsApp. Status: {resp.status_code}")
                            
                            return resumo
                        except Exception as e_net:
                            log_func(f"⚠️ Erro de rede ao enviar WhatsApp: {e_net}")
                            return resumo
                            
                    except Exception as e_rel:
                        log_func(f"⚠️ Não consegui gerar o relatório automático: {e_rel}")
                        return "Falha na geração do relatório."
                        
                except PermissionError:
                    log_func('❌ ERRO: Feche o Excel antes de salvar!')
                    return "Erro: Excel estava aberto."
            else:
                log_func("⚠️ Nenhuma foto foi injetada. Verifique se os nomes das pastas batem com as abas da planilha.")
                return "Nenhuma foto injetada."

            wb.close()
            del wb
            gc.collect()

        except Exception as e:
            if self.autonomous_mode:
                log_func("⚠️ [AUTO-EVOLUÇÃO] Falha detectada! Iniciando Code-Inspector no Micro-Kernel...")
                import traceback
                from core.auto_evolver import AutoEvolver
                evolver = AutoEvolver(log_func)
                patch_path = evolver.acionar_protocolo_mutacao(traceback.format_exc())
                if patch_path:
                    log_func(f"🛑 [HOT-RELOAD BLOQUEADO] A mutação foi bem sucedida no Sandbox.\nVerifique o patch gerado em: {patch_path}\nAplique-o manualmente para evitar corrupção e reinicie o aplicativo.")
                    return "Falha abortada de forma segura pelo Sandbox. Aguardando revisão manual do patch."
                else:
                    log_func("❌ [AUTO-EVOLUÇÃO] A auto-cura falhou ou não pôde ser validada de forma estável.")
            
            log_func(f"❌ Erro Crítico no motor: {e}")
            gc.collect()

    def auditar_efetivo(self, pasta_efetivo: str, excel_path: str, log_func):
        try:
            pastas_fisicas = [FileHandler.normalize_string(p.name) for p in Path(pasta_efetivo).iterdir()]
            
            wb = load_workbook(excel_path, read_only=True, data_only=True)
            ws = wb.active
            headers = [str(cell.value) for cell in ws[1] if cell.value is not None]
            
            colunas_possiveis = [i for i, c in enumerate(headers) if any(k in c.lower() for k in ['nome', 'funcionario', 'equipe', 'colaborador'])]
            col_idx = colunas_possiveis[0] if colunas_possiveis else 0
            
            nomes_excel_raw = []
            for row in ws.iter_rows(min_row=2, max_col=col_idx+1):
                val = row[col_idx].value
                if val:
                    nomes_excel_raw.append(str(val))
            
            wb.close()
            
            nomes_excel = [FileHandler.normalize_string(x) for x in set(nomes_excel_raw)]
            gc.collect()

            faltam_no_excel = []
            faltam_na_pasta = []

            def contains_fuzzy(nome, lista):
                for n in lista:
                    if nome in n or n in nome or difflib.SequenceMatcher(None, nome, n).ratio() > 0.8:
                        return True
                return False

            for p in pastas_fisicas:
                if not contains_fuzzy(p, nomes_excel):
                    faltam_no_excel.append(p.title())
            for n in nomes_excel:
                if not contains_fuzzy(n, pastas_fisicas) and len(n) > 2:
                    faltam_na_pasta.append(n.title())

            log_func("<br><b>=== RESULTADO DA AUDITORIA ===</b><br>")
            if faltam_no_excel:
                log_func("⚠️ <b>Na Pasta, mas NÃO no Excel:</b><br>" + "<br>".join(faltam_no_excel) + "<br>")
            if faltam_na_pasta:
                log_func("❌ <b>No Excel, mas NÃO mandou a pasta:</b><br>" + "<br>".join(faltam_na_pasta) + "<br>")
            if not faltam_no_excel and not faltam_na_pasta:
                log_func("✅ <b>Tudo 100% batendo! Nenhum GAP encontrado.</b>")
        except Exception as e:
            gc.collect()
            log_func(f"❌ Erro na auditoria: {e}")