import pandas as pd
from pathlib import Path
from core.file_handler import FileHandler
import gc
import unicodedata
import re
from openpyxl import load_workbook

class ExcelEngine:
    @staticmethod
    def validate_main_columns(excel_path: Path, expected_columns: list) -> bool:
        """Verifica de forma leve (read-only) se as colunas esperadas existem."""
        try:
            df = pd.read_excel(excel_path, nrows=0, engine='openpyxl')
            found_cols = [FileHandler.normalize_string(str(col)) for col in df.columns]
            norm_expected = [FileHandler.normalize_string(col) for col in expected_columns]
            del df
            gc.collect()
            return any(expected in found_cols for expected in norm_expected)
        except Exception as e:
            with open("erros_conhecidos.txt", "a", encoding="utf-8") as f:
                f.write(f"ExcelEngine: Falha na validacao {excel_path.name} - {e}\n")
            return False

    @staticmethod
    def inject_formula(excel_path: str, sheet_name: str, cell_coord: str, formula: str) -> bool:
        """Injeta uma fórmula nativa no Excel utilizando openpyxl para máxima compatibilidade."""
        try:
            wb = load_workbook(excel_path)
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                ws[cell_coord] = formula
                wb.save(excel_path)
                wb.close()
                del wb
                gc.collect()
                return True
            return False
        except Exception as e:
            with open("erros_conhecidos.txt", "a", encoding="utf-8") as f:
                f.write(f"ExcelEngine: Erro ao injetar formula em {excel_path} - {e}\n")
            return False

    @staticmethod
    def _norm(text: str) -> str:
        """Normaliza texto para comparação: lowercase, sem acentos, sem espaços extras."""
        if not isinstance(text, str): text = str(text)
        text = text.strip().lower()
        text = ''.join(c for c in unicodedata.normalize('NFD', text) if unicodedata.category(c) != 'Mn')
        return re.sub(r'\s+', ' ', text)

    @staticmethod
    def query_data(excel_path: str, alvo: str, coluna_desejada: str, sheet_name: str = None) -> str:
        try:
            import difflib
            path = Path(excel_path)
            if not path.exists(): return f"❌ Arquivo não encontrado: {excel_path}"
            df = pd.read_excel(excel_path, sheet_name=sheet_name if sheet_name else 0, engine='openpyxl', dtype=str)
            norm_cols = {ExcelEngine._norm(col): col for col in df.columns}
            alvo_norm = ExcelEngine._norm(alvo)
            coluna_norm = ExcelEngine._norm(coluna_desejada)
            coluna_real = None
            for nc, orig in norm_cols.items():
                if coluna_norm in nc or nc in coluna_norm or difflib.SequenceMatcher(None, coluna_norm, nc).ratio() > 0.8:
                    coluna_real = orig
                    break
            if not coluna_real:
                del df; gc.collect()
                return f"⚠️ Coluna '{coluna_desejada}' não localizada."
            def is_fuzzy(val):
                v = ExcelEngine._norm(str(val))
                if alvo_norm in v or v in alvo_norm: return True
                return difflib.SequenceMatcher(None, alvo_norm, v).ratio() > 0.8
            mask = df.apply(lambda col: col.astype(str).apply(is_fuzzy), axis=0).any(axis=1)
            resultado = df[mask]
            if resultado.empty:
                del df; gc.collect()
                return f"🔍 Nada encontrado para '{alvo}'."
            valores = resultado[coluna_real].dropna().unique().tolist()
            resposta = ", ".join(str(v) for v in valores)
            del df; gc.collect()
            return f"✅ '{alvo}' → {coluna_desejada.capitalize()}: **{resposta}**"
        except Exception as e:
            gc.collect()
            return f"❌ Erro na consulta: {e}"

    @staticmethod
    def query_count_empty(excel_path: str, coluna: str, sheet_name: str = None) -> str:
        try:
            path = Path(excel_path)
            if not path.exists(): return f"❌ Arquivo não encontrado: {excel_path}"
            df = pd.read_excel(excel_path, sheet_name=sheet_name if sheet_name else 0, engine='openpyxl', dtype=str)
            norm_cols = {ExcelEngine._norm(col): col for col in df.columns}
            coluna_norm = ExcelEngine._norm(coluna)
            coluna_real = None
            for nc, orig in norm_cols.items():
                if coluna_norm in nc or nc in coluna_norm:
                    coluna_real = orig
                    break
            if coluna_real is None:
                del df; gc.collect()
                return f"⚠️ Coluna '{coluna}' não localizada na planilha."
            vazias = df[coluna_real].apply(lambda x: str(x).strip() in ['', 'nan', 'None', 'NaN']).sum()
            del df; gc.collect()
            return f"📊 Encontrei **{int(vazias)} registro(s)** sem valor na coluna '{coluna_real}'."
        except Exception as e:
            gc.collect()
            return f"❌ Erro ao contar registros: {e}"
