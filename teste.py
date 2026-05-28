import os
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import CellIsRule

def criar_planilha():
    try:
        # Definir o caminho de destino da planilha (priorizando OneDrive/Desktop)
        desktop_path = Path.home() / 'OneDrive' / 'Desktop'
        if not desktop_path.exists():
            desktop_path = Path.home() / 'Desktop'
            
        planilha_path = desktop_path / 'teeste' / 'controle_energia.xlsx'

        # Criar o diretório de destino se não existir
        os.makedirs(planilha_path.parent, exist_ok=True)

        # Criar a planilha
        wb = Workbook()
        ws = wb.active

        # Definir o título da planilha e habilitar as linhas de grade
        ws.title = 'Controle de Energia'
        ws.views.sheetView[0].showGridLines = True

        # Estilos de borda
        thin_border = Border(
            left=Side(style='thin', color='D3D3D3'),
            right=Side(style='thin', color='D3D3D3'),
            top=Side(style='thin', color='D3D3D3'),
            bottom=Side(style='thin', color='D3D3D3')
        )
        double_bottom_border = Border(
            top=Side(style='thin', color='000000'),
            bottom=Side(style='double', color='000000'),
            left=Side(style='thin', color='D3D3D3'),
            right=Side(style='thin', color='D3D3D3')
        )

        # 1. Cabeçalho Principal (Título)
        ws.merge_cells('A1:I1')
        ws['A1'] = 'SISTEMA DE CONTROLE DE ENERGIA - CONDOMÍNIO GD'
        ws['A1'].font = Font(name='Segoe UI', size=16, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color='003366', end_color='003366', fill_type='solid')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 45

        # 2. Seção A: Tabela de Resumo Mensal (Colunas A-D)
        ws.merge_cells('A3:D3')
        ws['A3'] = 'Resumo Mensal de Consumo'
        ws['A3'].font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
        ws['A3'].fill = PatternFill(start_color='005B96', end_color='005B96', fill_type='solid')
        ws['A3'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[3].height = 25

        headers_a = ["Mês", "Consumo Total (kWh)", "Custo Total (R$)", "Média / Unidade (R$)"]
        for col_idx, text in enumerate(headers_a, start=1):
            cell = ws.cell(row=4, column=col_idx, value=text)
            cell.font = Font(name='Segoe UI', size=9, bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='00838F', end_color='00838F', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.row_dimensions[4].height = 25

        # Dados da Tabela de Resumo
        dados_resumo = [
            ["Janeiro", 1200],
            ["Fevereiro", 1350],
            ["Março", 1100],
            ["Abril", 1400],
            ["Maio", 1500],
            ["Junho", 1650],
        ]

        for i, (mes, consumo) in enumerate(dados_resumo, start=5):
            # Mês
            cell_mes = ws.cell(row=i, column=1, value=mes)
            cell_mes.font = Font(name='Segoe UI', size=10)
            cell_mes.alignment = Alignment(horizontal='center', vertical='center')
            cell_mes.border = thin_border
            
            # Consumo
            cell_cons = ws.cell(row=i, column=2, value=consumo)
            cell_cons.font = Font(name='Segoe UI', size=10)
            cell_cons.alignment = Alignment(horizontal='right', vertical='center')
            cell_cons.number_format = '#,##0'
            cell_cons.border = thin_border
            
            # Custo (Consumo * R$ 0.65)
            cell_custo = ws.cell(row=i, column=3, value=f"=B{i}*0.65")
            cell_custo.font = Font(name='Segoe UI', size=10)
            cell_custo.alignment = Alignment(horizontal='right', vertical='center')
            cell_custo.number_format = 'R$ #,##0.00'
            cell_custo.border = thin_border
            
            # Média por Unidade (Custo / 10 unidades)
            cell_media = ws.cell(row=i, column=4, value=f"=C{i}/10")
            cell_media.font = Font(name='Segoe UI', size=10)
            cell_media.alignment = Alignment(horizontal='right', vertical='center')
            cell_media.number_format = 'R$ #,##0.00'
            cell_media.border = thin_border

            # Zebra Striping
            if i % 2 == 0:
                bg_color = 'F2F9F9'
                cell_mes.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
                cell_cons.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
                cell_custo.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
                cell_media.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
            
            ws.row_dimensions[i].height = 20

        # Linha de Totais da Tabela de Resumo (Row 11)
        totais_a = [
            ("Total/Média", Alignment(horizontal='center', vertical='center')),
            ("=SUM(B5:B10)", Alignment(horizontal='right', vertical='center')),
            ("=SUM(C5:C10)", Alignment(horizontal='right', vertical='center')),
            ("=AVERAGE(D5:D10)", Alignment(horizontal='right', vertical='center'))
        ]
        for col_idx, (formula, align) in enumerate(totais_a, start=1):
            cell = ws.cell(row=11, column=col_idx, value=formula)
            cell.font = Font(name='Segoe UI', size=10, bold=True)
            cell.fill = PatternFill(start_color='E0F7FA', end_color='E0F7FA', fill_type='solid')
            cell.alignment = align
            cell.border = double_bottom_border
            if col_idx in [2]:
                cell.number_format = '#,##0'
            elif col_idx in [3, 4]:
                cell.number_format = 'R$ #,##0.00'
        ws.row_dimensions[11].height = 22

        # 3. Seção B: Detalhamento por Unidade Consumidora (Colunas F-I)
        ws.merge_cells('F3:I3')
        ws['F3'] = 'Detalhamento por Unidade (Mês Atual)'
        ws['F3'].font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
        ws['F3'].fill = PatternFill(start_color='005B96', end_color='005B96', fill_type='solid')
        ws['F3'].alignment = Alignment(horizontal='center', vertical='center')

        headers_b = ["Unidade", "Consumo (kWh)", "Valor (R$)", "Status de Consumo"]
        for col_idx, text in enumerate(headers_b, start=6):
            cell = ws.cell(row=4, column=col_idx, value=text)
            cell.font = Font(name='Segoe UI', size=9, bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='00838F', end_color='00838F', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        dados_detalhados = [
            ["Apto 101", 150],
            ["Apto 102", 180],
            ["Apto 103", 90],
            ["Apto 104", 220],
            ["Apto 201", 130],
            ["Apto 202", 140],
            ["Apto 203", 210],
            ["Apto 204", 160],
            ["Apto 205", 110],
            ["Apto 206", 150],
        ]

        for i, (unidade, consumo) in enumerate(dados_detalhados, start=5):
            # Unidade
            cell_un = ws.cell(row=i, column=6, value=unidade)
            cell_un.font = Font(name='Segoe UI', size=10)
            cell_un.alignment = Alignment(horizontal='center', vertical='center')
            cell_un.border = thin_border
            
            # Consumo (desbloqueado para edição do usuário)
            cell_cons = ws.cell(row=i, column=7, value=consumo)
            cell_cons.font = Font(name='Segoe UI', size=10)
            cell_cons.alignment = Alignment(horizontal='right', vertical='center')
            cell_cons.number_format = '#,##0'
            cell_cons.border = thin_border
            cell_cons.protection = Protection(locked=False)
            
            # Valor (Consumo * R$ 0.65)
            cell_valor = ws.cell(row=i, column=8, value=f"=G{i}*0.65")
            cell_valor.font = Font(name='Segoe UI', size=10)
            cell_valor.alignment = Alignment(horizontal='right', vertical='center')
            cell_valor.number_format = 'R$ #,##0.00'
            cell_valor.border = thin_border
            
            # Status de Consumo (Fórmula condicional baseada no limite de 160 kWh)
            cell_status = ws.cell(row=i, column=9, value=f'=IF(G{i}>160,"Alto","Normal")')
            cell_status.font = Font(name='Segoe UI', size=10, bold=True)
            cell_status.alignment = Alignment(horizontal='center', vertical='center')
            cell_status.border = thin_border

            # Zebra Striping
            if i % 2 == 0:
                bg_color = 'F2F9F9'
                cell_un.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
                cell_cons.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
                cell_valor.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
                cell_status.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')

        # Linha de Totais da Tabela de Detalhes (Row 15)
        totais_b = [
            ("Total/Média", Alignment(horizontal='center', vertical='center')),
            ("=SUM(G5:G14)", Alignment(horizontal='right', vertical='center')),
            ("=SUM(H5:H14)", Alignment(horizontal='right', vertical='center')),
            ("", Alignment(horizontal='center', vertical='center'))
        ]
        for col_idx, (formula, align) in enumerate(totais_b, start=6):
            cell = ws.cell(row=15, column=col_idx, value=formula)
            cell.font = Font(name='Segoe UI', size=10, bold=True)
            cell.fill = PatternFill(start_color='E0F7FA', end_color='E0F7FA', fill_type='solid')
            cell.alignment = align
            cell.border = double_bottom_border
            if col_idx in [7]:
                cell.number_format = '#,##0'
            elif col_idx in [8]:
                cell.number_format = 'R$ #,##0.00'
        ws.row_dimensions[15].height = 22

        # 4. Formatação Condicional Nativa do Excel no Status
        red_fill = PatternFill(start_color='FFCDD2', end_color='FFCDD2', fill_type='solid')
        red_font = Font(name='Segoe UI', size=10, bold=True, color='B71C1C')
        green_fill = PatternFill(start_color='C8E6C9', end_color='C8E6C9', fill_type='solid')
        green_font = Font(name='Segoe UI', size=10, bold=True, color='1B5E20')

        ws.conditional_formatting.add('I5:I14', CellIsRule(operator='equal', formula=['"Alto"'], stopIfTrue=True, fill=red_fill, font=red_font))
        ws.conditional_formatting.add('I5:I14', CellIsRule(operator='equal', formula=['"Normal"'], stopIfTrue=True, fill=green_fill, font=green_font))

        # 5. Legenda de Consumo (Abaixo da tabela detalhada)
        ws.merge_cells('F17:I17')
        ws['F17'] = 'Legenda de Status'
        ws['F17'].font = Font(name='Segoe UI', size=10, bold=True, color='FFFFFF')
        ws['F17'].fill = PatternFill(start_color='00838F', end_color='00838F', fill_type='solid')
        ws['F17'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[17].height = 20

        # Legenda Alto
        ws.merge_cells('F18:G18')
        ws['F18'] = 'Consumo Alto (> 160 kWh)'
        ws['F18'].font = Font(name='Segoe UI', size=9, bold=True, color='B71C1C')
        ws['F18'].fill = PatternFill(start_color='FFCDD2', end_color='FFCDD2', fill_type='solid')
        ws['F18'].alignment = Alignment(horizontal='center', vertical='center')
        ws['F18'].border = thin_border
        
        # Legenda Normal
        ws.merge_cells('H18:I18')
        ws['H18'] = 'Consumo Normal (≤ 160 kWh)'
        ws['H18'].font = Font(name='Segoe UI', size=9, bold=True, color='1B5E20')
        ws['H18'].fill = PatternFill(start_color='C8E6C9', end_color='C8E6C9', fill_type='solid')
        ws['H18'].alignment = Alignment(horizontal='center', vertical='center')
        ws['H18'].border = thin_border
        ws.row_dimensions[18].height = 20

        # 6. Botão Interativo Simulado para Atualizar Fórmulas
        ws.merge_cells('F20:I21')
        ws['F20'] = '🔄 ATUALIZAR DADOS (F9 / Recalcular)'
        ws['F20'].font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
        ws['F20'].fill = PatternFill(start_color='00838F', end_color='00838F', fill_type='solid')
        ws['F20'].alignment = Alignment(horizontal='center', vertical='center')
        
        button_border = Border(
            left=Side(style='medium', color='004D40'),
            right=Side(style='medium', color='004D40'),
            top=Side(style='medium', color='004D40'),
            bottom=Side(style='medium', color='004D40')
        )
        for r in range(20, 22):
            for c in range(6, 10):
                cell = ws.cell(row=r, column=c)
                cell.border = button_border

        # 7. Gráfico de Consumo Mensal (Abaixo da tabela resumo)
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Histórico de Consumo de Energia Mensal"
        chart.y_axis.title = "Consumo (kWh)"
        chart.x_axis.title = "Mês"
        chart.width = 12
        chart.height = 8.5
        
        data = Reference(ws, min_col=2, min_row=4, max_row=10)
        cats = Reference(ws, min_col=1, min_row=5, max_row=10)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.legend = None
        
        ws.add_chart(chart, "A13")

        # 8. Observações Gerais
        ws.merge_cells('F23:I23')
        ws['F23'] = 'Observações / Notas'
        ws['F23'].font = Font(name='Segoe UI', size=10, bold=True)
        ws['F23'].fill = PatternFill(start_color='E0F7FA', end_color='E0F7FA', fill_type='solid')
        ws['F23'].alignment = Alignment(horizontal='left', vertical='center')
        ws['F23'].border = thin_border
        
        ws.merge_cells('F24:I26')
        ws['F24'] = (
            "1. A tarifa atual de energia considerada é de R$ 0,65 por kWh.\n"
            "2. As células de consumo das unidades (coluna G) estão DESBLOQUEADAS.\n"
            "3. O sistema de proteção de planilha impede alterações indesejadas em fórmulas."
        )
        ws['F24'].font = Font(name='Segoe UI', size=9, italic=True)
        ws['F24'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        border_obs = Border(
            left=Side(style='thin', color='D3D3D3'),
            right=Side(style='thin', color='D3D3D3'),
            top=Side(style='thin', color='D3D3D3'),
            bottom=Side(style='thin', color='D3D3D3')
        )
        for r in range(24, 27):
            for c in range(6, 10):
                ws.cell(row=r, column=c).border = border_obs

        # 9. Proteção da Planilha
        ws.protection.sheet = True
        ws.protection.enable()

        # 10. Ajuste Automático da Largura das Colunas
        from openpyxl.utils import get_column_letter
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                # Se a célula for mesclada, pula o cálculo da largura baseada nela para evitar distorção
                coord = cell.coordinate
                is_merged = False
                for merged_range in ws.merged_cells.ranges:
                    if coord in merged_range and coord != merged_range.start_cell.coordinate:
                        is_merged = True
                        break
                if is_merged:
                    continue
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            # Define uma largura mínima razoável
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

        # Salvar a planilha
        wb.save(planilha_path)
        print(f'Planilha premium criada com sucesso em: {planilha_path}')

    except Exception as e:
        print(f'Erro ao criar planilha: {e}')

if __name__ == '__main__':
    criar_planilha()